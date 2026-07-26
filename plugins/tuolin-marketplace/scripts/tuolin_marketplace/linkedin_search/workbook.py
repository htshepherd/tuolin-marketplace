from __future__ import annotations

import hashlib
import json
import os
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from .browser_contract import normalize_linkedin_url


CONTACT_SHEET = "潜客联系人"
EVIDENCE_SHEET = "潜客证据"
DISPATCH_SHEET = "发送记录"
CONTACT_HEADERS = (
    "联系人ID", "LinkedIn主页", "姓名", "职位", "公司", "公司主页", "连接状态",
    "首次发现时间", "最近发现时间", "老板判断", "老板备注",
)
EVIDENCE_HEADERS = (
    "证据ID", "联系人ID", "运行ID", "关键词", "完整贴文", "贴文链接", "Codex初步判断",
    "商业角色", "支持证据", "不确定性", "Codex建议", "发现时间",
)
DISPATCH_HEADERS = (
    "批次摘要", "联系人ID", "授权时间", "留言", "间隔秒数", "尝试时间", "结果", "结果时间",
)
DECISIONS = ("发送", "排除", "待定")


def prospect_workbook_path(run_dir: Path, account_profile_url: str) -> Path:
    account = normalize_linkedin_url(account_profile_url)
    key = hashlib.sha256(account.encode("utf-8")).hexdigest()[:20]
    shared = Path(run_dir).expanduser().resolve().parent / "shared" / "prospect-workbooks"
    shared.mkdir(parents=True, exist_ok=True)
    return shared / f"linkedin-prospects-{key}.xlsx"


def load_contact_index(run_dir: Path, account_profile_url: str) -> dict[str, dict[str, Any]]:
    path = prospect_workbook_path(run_dir, account_profile_url)
    if not path.exists():
        return {}
    workbook = _load_valid_workbook(path)
    sheet = workbook[CONTACT_SHEET]
    headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1)}
    result: dict[str, dict[str, Any]] = {}
    for row in range(2, sheet.max_row + 1):
        profile = sheet.cell(row, headers["LinkedIn主页"]).value
        if not profile:
            continue
        normalized = normalize_linkedin_url(str(profile))
        result[normalized] = {
            header: sheet.cell(row, column).value for header, column in headers.items()
        } | {"row_number": row}
    workbook.close()
    return result


def sync_candidate(
    run_dir: Path,
    *,
    account_profile_url: str,
    card: dict[str, Any],
    now: datetime | None = None,
) -> dict[str, Any]:
    path = prospect_workbook_path(run_dir, account_profile_url)
    workbook = _load_valid_workbook(path) if path.exists() else _new_workbook(account_profile_url)
    contact_sheet = workbook[CONTACT_SHEET]
    evidence_sheet = workbook[EVIDENCE_SHEET]
    profile = normalize_linkedin_url(str(card["selected_member"]["profile_url"]))
    company_url = normalize_linkedin_url(str((card.get("company") or {}).get("url") or "")) if (card.get("company") or {}).get("url") else ""
    contact_id = "contact_" + hashlib.sha256(profile.encode("utf-8")).hexdigest()[:16]
    timestamp = _iso(now)
    row = _find_contact_row(contact_sheet, profile)
    is_new = row is None
    if is_new:
        row = contact_sheet.max_row + 1
        values = (
            contact_id, profile, card["selected_member"].get("name", ""), card["selected_member"].get("title", ""),
            (card.get("company") or {}).get("name", ""), company_url, "none", timestamp, timestamp, "待定", "",
        )
        for column, value in enumerate(values, start=1):
            contact_sheet.cell(row, column, value)
    else:
        contact_sheet.cell(row, CONTACT_HEADERS.index("最近发现时间") + 1, timestamp)
    evidence_seed = f"{run_dir.resolve().name}|{card.get('post_url')}|{contact_id}"
    evidence_id = "evidence_" + hashlib.sha256(evidence_seed.encode("utf-8")).hexdigest()[:20]
    evidence_exists = any(evidence_sheet.cell(index, 1).value == evidence_id for index in range(2, evidence_sheet.max_row + 1))
    if not evidence_exists:
        assessment = card.get("preliminary_assessment") or {}
        values = (
            evidence_id,
            contact_id,
            run_dir.resolve().name,
            card.get("source_keyword", ""),
            card.get("post_text", ""),
            card.get("post_url", ""),
            assessment.get("summary") or card.get("relevance_reason", ""),
            assessment.get("business_role") or card.get("business_role", ""),
            assessment.get("supporting_evidence") or card.get("supporting_evidence", ""),
            assessment.get("uncertainty") or card.get("material_doubts", ""),
            assessment.get("recommendation") or "保留给老板判断",
            timestamp,
        )
        evidence_sheet.append(values)
    previous_revision = _revision(workbook)
    revision = previous_revision + 1
    workbook.properties.subject = f"tuolin-linkedin-account={normalize_linkedin_url(account_profile_url)}"
    workbook.properties.description = f"revision={revision}"
    _save_atomic(workbook, path)
    workbook.close()
    receipt = {
        "schema_version": 1,
        "workbook_path": str(path),
        "account_profile_url": normalize_linkedin_url(account_profile_url),
        "pre_revision": previous_revision,
        "post_revision": revision,
        "contact_ids": [contact_id] if is_new else [],
        "updated_contact_ids": [] if is_new else [contact_id],
        "evidence_ids": [] if evidence_exists else [evidence_id],
        "digest": _file_digest(path),
        "synced_at": timestamp,
    }
    receipt_dir = Path(run_dir).resolve() / "workbook-sync"
    receipt_dir.mkdir(parents=True, exist_ok=True)
    receipt_path = receipt_dir / f"revision-{revision:06d}.json"
    _write_json_atomic(receipt_path, receipt)
    return {**receipt, "receipt_path": str(receipt_path), "is_new_contact": is_new, "contact_id": contact_id, "evidence_id": evidence_id}


def read_current_selections(run_dir: Path, account_profile_url: str) -> list[dict[str, Any]]:
    path = prospect_workbook_path(run_dir, account_profile_url)
    if not path.exists():
        return []
    workbook = _load_valid_workbook(path)
    sheet = workbook[CONTACT_SHEET]
    headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1)}
    evidence_sheet = workbook[EVIDENCE_SHEET]
    evidence_headers = {cell.value: index for index, cell in enumerate(evidence_sheet[1], start=1)}
    latest_evidence: dict[str, dict[str, Any]] = {}
    for evidence_row in range(2, evidence_sheet.max_row + 1):
        contact_id = str(evidence_sheet.cell(evidence_row, evidence_headers["联系人ID"]).value or "")
        if contact_id:
            latest_evidence[contact_id] = {
                header: evidence_sheet.cell(evidence_row, column).value for header, column in evidence_headers.items()
            }
    selected = []
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, headers["老板判断"]).value != "发送":
            continue
        record = {header: sheet.cell(row, column).value for header, column in headers.items()}
        if str(record.get("连接状态") or "none").casefold() not in {"none", "failed_no_dispatch"}:
            continue
        record["最新证据"] = latest_evidence.get(str(record.get("联系人ID") or ""), {})
        selected.append(record)
    workbook.close()
    return selected


def read_all_contacts(run_dir: Path, account_profile_url: str) -> list[dict[str, Any]]:
    path = prospect_workbook_path(run_dir, account_profile_url)
    if not path.exists():
        return []
    workbook = _load_valid_workbook(path)
    sheet = workbook[CONTACT_SHEET]
    headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1)}
    records = [
        {header: sheet.cell(row, column).value for header, column in headers.items()}
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, headers["联系人ID"]).value
    ]
    workbook.close()
    return records


def set_boss_decision(
    run_dir: Path,
    account_profile_url: str,
    profile_url: str,
    decision: str,
    note: str = "",
) -> Path:
    if decision not in DECISIONS:
        raise ValueError("老板判断只允许发送、排除或待定。")
    path = prospect_workbook_path(run_dir, account_profile_url)
    workbook = _load_valid_workbook(path)
    sheet = workbook[CONTACT_SHEET]
    row = _find_contact_row(sheet, normalize_linkedin_url(profile_url))
    if row is None:
        workbook.close()
        raise ValueError("累计潜客表中没有该联系人。")
    sheet.cell(row, CONTACT_HEADERS.index("老板判断") + 1, decision)
    sheet.cell(row, CONTACT_HEADERS.index("老板备注") + 1, note)
    workbook.properties.description = f"revision={_revision(workbook) + 1}"
    _save_atomic(workbook, path)
    workbook.close()
    return path


def append_dispatch_record(
    run_dir: Path,
    *,
    account_profile_url: str,
    batch_digest: str,
    contact_id: str,
    authorized_at: str = "",
    note: str = "",
    interval_seconds: int = 0,
    attempted_at: str = "",
    result: str = "authorized",
    result_at: str = "",
) -> Path:
    return append_dispatch_records(
        run_dir,
        account_profile_url=account_profile_url,
        records=[{
            "batch_digest": batch_digest,
            "contact_id": contact_id,
            "authorized_at": authorized_at,
            "note": note,
            "interval_seconds": interval_seconds,
            "attempted_at": attempted_at,
            "result": result,
            "result_at": result_at,
        }],
    )


def append_dispatch_records(
    run_dir: Path,
    *,
    account_profile_url: str,
    records: list[dict[str, Any]],
) -> Path:
    if not records:
        raise ValueError("至少需要一条发送历史记录。")
    path = prospect_workbook_path(run_dir, account_profile_url)
    workbook = _load_valid_workbook(path)
    for record in records:
        workbook[DISPATCH_SHEET].append((
            record.get("batch_digest", ""),
            record.get("contact_id", ""),
            record.get("authorized_at", ""),
            record.get("note", ""),
            int(record.get("interval_seconds") or 0),
            record.get("attempted_at", ""),
            record.get("result", ""),
            record.get("result_at", ""),
        ))
    workbook.properties.description = f"revision={_revision(workbook) + 1}"
    _save_atomic(workbook, path)
    workbook.close()
    return path


def update_contact_state(run_dir: Path, account_profile_url: str, profile_url: str, state: str) -> Path:
    path = prospect_workbook_path(run_dir, account_profile_url)
    workbook = _load_valid_workbook(path)
    sheet = workbook[CONTACT_SHEET]
    row = _find_contact_row(sheet, normalize_linkedin_url(profile_url))
    if row is None:
        workbook.close()
        raise ValueError("累计潜客表中没有该联系人。")
    sheet.cell(row, CONTACT_HEADERS.index("连接状态") + 1, state)
    workbook.properties.description = f"revision={_revision(workbook) + 1}"
    _save_atomic(workbook, path)
    workbook.close()
    return path


def _new_workbook(account_profile_url: str):
    workbook = Workbook()
    contact = workbook.active
    contact.title = CONTACT_SHEET
    evidence = workbook.create_sheet(EVIDENCE_SHEET)
    dispatch = workbook.create_sheet(DISPATCH_SHEET)
    contact.append(CONTACT_HEADERS)
    evidence.append(EVIDENCE_HEADERS)
    dispatch.append(DISPATCH_HEADERS)
    validation = DataValidation(type="list", formula1='"发送,排除,待定"', allow_blank=False)
    contact.add_data_validation(validation)
    validation.add("J2:J1048576")
    contact.freeze_panes = "A2"
    evidence.freeze_panes = "A2"
    dispatch.freeze_panes = "A2"
    workbook.properties.subject = f"tuolin-linkedin-account={normalize_linkedin_url(account_profile_url)}"
    workbook.properties.description = "revision=0"
    return workbook


def _load_valid_workbook(path: Path):
    try:
        workbook = load_workbook(path)
    except Exception as exc:
        raise ValueError(f"累计潜客工作簿无法读取或已损坏：{path}：{exc}") from exc
    if workbook.sheetnames != [CONTACT_SHEET, EVIDENCE_SHEET, DISPATCH_SHEET]:
        workbook.close()
        raise ValueError("累计潜客工作簿结构无效；必须保留三个稳定工作表。")
    expected = {CONTACT_SHEET: CONTACT_HEADERS, EVIDENCE_SHEET: EVIDENCE_HEADERS, DISPATCH_SHEET: DISPATCH_HEADERS}
    for name, headers in expected.items():
        actual = tuple(cell.value for cell in workbook[name][1])
        if actual != headers:
            workbook.close()
            raise ValueError(f"累计潜客工作簿 {name} 表头无效。")
    return workbook


def _find_contact_row(sheet, profile_url: str) -> int | None:
    profile_column = CONTACT_HEADERS.index("LinkedIn主页") + 1
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, profile_column).value
        if value and normalize_linkedin_url(str(value)) == profile_url:
            return row
    return None


def _revision(workbook) -> int:
    description = str(workbook.properties.description or "")
    if description.startswith("revision="):
        try:
            return int(description.split("=", 1)[1])
        except ValueError:
            pass
    return 0


def _save_atomic(workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(prefix=path.stem + "-", suffix=".xlsx", dir=path.parent)
    os.close(descriptor)
    temporary = Path(temporary_name)
    backup = path.with_suffix(".previous.xlsx")
    try:
        workbook.save(temporary)
        check = load_workbook(temporary, read_only=True)
        check.close()
        if path.exists():
            shutil.copy2(path, backup)
        temporary.replace(path)
    except Exception as exc:
        if temporary.exists():
            temporary.unlink()
        raise ValueError(f"累计潜客工作簿写入失败；最后好版本保持不变：{path}：{exc}") from exc


def _file_digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_json_atomic(path: Path, value: dict[str, Any]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def _iso(now: datetime | None) -> str:
    value = now or datetime.now().astimezone()
    if value.tzinfo is None:
        value = value.astimezone()
    return value.isoformat()
