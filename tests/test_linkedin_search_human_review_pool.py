from __future__ import annotations

import json
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from scripts.tuolin_marketplace.linkedin_search.agent import create_linkedin_search_run
from scripts.tuolin_marketplace.linkedin_search.browser_contract import (
    LinkedInAccountObservation,
    LinkedInPostSearchObservation,
    bind_linkedin_account,
    finish_current_keyword,
    record_first_posts_search,
)
from scripts.tuolin_marketplace.linkedin_search.feedback import (
    capture_workbook_feedback,
    confirm_screening_rule,
    propose_screening_rule,
)
from scripts.tuolin_marketplace.linkedin_search.interview import build_search_interview, confirmed_search_brief
from scripts.tuolin_marketplace.linkedin_search.ledger import ledger_path
from scripts.tuolin_marketplace.linkedin_search.review import (
    confirm_candidate_batch,
    prepare_candidate_batch_review,
    prepare_dispatch_authorization,
)
from scripts.tuolin_marketplace.linkedin_search.review_pool import (
    allocate_first_pass_shares,
    finish_keyword,
    initialize_review_pool,
    register_contact,
)
from scripts.tuolin_marketplace.linkedin_search.workbook import (
    CONTACT_SHEET,
    DECISIONS,
    DISPATCH_SHEET,
    EVIDENCE_SHEET,
    load_contact_index,
    prospect_workbook_path,
    set_boss_decision,
    sync_candidate,
)
from scripts.tuolin_marketplace.project_layout import resolve_paths
from tests.test_linkedin_search_agent import _record_demo_candidate


ACCOUNT = "https://www.linkedin.com/in/tuolin-sales"


class HumanReviewPoolTests(unittest.TestCase):
    def test_interview_uses_review_pool_and_integer_minute_interval(self) -> None:
        interview = build_search_interview(
            "LinkedIn 搜索，关键词：Exhaust Wrap, Exhaust Heat Wrap；排序最新；近一个月；"
            "不使用留言；间隔 2 分钟；最多找 50 人"
        )
        self.assertTrue(interview["completed"])
        brief = confirmed_search_brief(interview)
        self.assertEqual(brief["human_review_pool_limit"], 50)
        self.assertEqual(brief["interval_seconds"], 120)
        self.assertNotIn("requested_limit", brief)
        self.assertNotIn("opened_post_limit_per_keyword", brief)
        with self.assertRaisesRegex(ValueError, "1–100"):
            build_search_interview(
                "LinkedIn 搜索，关键词：Wrap；排序最新；近一个月；不使用留言；间隔 2 分钟；最多找 101 人"
            )
        with self.assertRaisesRegex(ValueError, "整数分钟"):
            build_search_interview(
                "LinkedIn 搜索，关键词：Wrap；排序最新；近一个月；不使用留言；间隔 0 分钟；最多找 50 人"
            )

    def test_balanced_keyword_sampling_is_deterministic_and_refills(self) -> None:
        self.assertEqual(allocate_first_pass_shares(10, ["a", "b", "c"]), [4, 3, 3])
        pool = initialize_review_pool(4, ["a", "b"])
        pool = register_contact(pool, "a1", is_new=True)
        pool = register_contact(pool, "a2", is_new=True)
        pool, reason, next_keyword = finish_keyword(pool)
        self.assertEqual(reason, "first_pass_soft_share_reached")
        self.assertEqual(next_keyword, "b")
        pool["keywords"][1]["consecutive_bottom_no_growth_cycles"] = 3
        pool, reason, next_keyword = finish_keyword(pool)
        self.assertEqual(next_keyword, "a")
        self.assertEqual(pool["pass"], "refill")
        pool = register_contact(pool, "a3", is_new=True)
        pool = register_contact(pool, "a4", is_new=True)
        pool, reason, next_keyword = finish_keyword(pool)
        self.assertEqual(reason, "human_review_pool_full")
        self.assertIsNone(next_keyword)

    def test_workbook_round_trip_dropdown_and_repeated_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            run_dir = Path(tmp) / "reports" / "linkedin-search" / "run-1"
            run_dir.mkdir(parents=True)
            card = _card("one", "First complete post")
            first = sync_candidate(run_dir, account_profile_url=ACCOUNT, card=card)
            self.assertTrue(first["is_new_contact"])
            card["post_url"] = "https://linkedin.com/posts/one-second"
            card["post_text"] = "Second complete post"
            second = sync_candidate(run_dir, account_profile_url=ACCOUNT, card=card)
            self.assertFalse(second["is_new_contact"])
            path = prospect_workbook_path(run_dir, ACCOUNT)
            workbook = load_workbook(path)
            self.assertEqual(workbook.sheetnames, [CONTACT_SHEET, EVIDENCE_SHEET, DISPATCH_SHEET])
            self.assertEqual(workbook[CONTACT_SHEET].max_row, 2)
            self.assertEqual(workbook[EVIDENCE_SHEET].max_row, 3)
            self.assertEqual(workbook[CONTACT_SHEET]["J2"].value, "待定")
            formulas = [validation.formula1 for validation in workbook[CONTACT_SHEET].data_validations.dataValidation]
            self.assertIn('"发送,排除,待定"', formulas)
            self.assertEqual(tuple(DECISIONS), ("发送", "排除", "待定"))
            workbook.close()

    def test_reduced_capacity_allows_discovery_but_twelve_selection_requires_exact_subset(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            paths = resolve_paths(Path(tmp), {})
            result = _create_bound_discovery(paths, pool_limit=12, successes=95)
            run_dir = Path(result.run_dir)
            state = json.loads((run_dir / "workflow_state.json").read_text(encoding="utf-8"))
            self.assertEqual(state["phase"], "discovering_posts")
            self.assertEqual(state["capacity_at_account_binding"]["remaining_capacity"], 5)
            for index in range(12):
                _record_demo_candidate(run_dir, suffix=f"pool-{index}")
            finish_current_keyword(run_dir)
            for profile in load_contact_index(run_dir, ACCOUNT):
                set_boss_decision(run_dir, ACCOUNT, profile, "发送")
            prepare_candidate_batch_review(run_dir)
            confirm_candidate_batch(run_dir)
            with self.assertRaisesRegex(ValueError, "不会自动截取前 N 人"):
                prepare_dispatch_authorization(run_dir)

    def test_zero_capacity_still_allows_workbook_discovery_but_blocks_connect_snapshot(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            paths = resolve_paths(Path(tmp), {})
            result = _create_bound_discovery(paths, pool_limit=1, successes=100)
            run_dir = Path(result.run_dir)
            _record_demo_candidate(run_dir, suffix="zero-capacity")
            finish_current_keyword(run_dir)
            set_boss_decision(run_dir, ACCOUNT, "https://linkedin.com/in/demo-zero-capacity", "发送")
            prepare_candidate_batch_review(run_dir)
            confirm_candidate_batch(run_dir)
            with self.assertRaisesRegex(ValueError, "容量为 0"):
                prepare_dispatch_authorization(run_dir)

    def test_twelve_selected_contacts_form_twelve_person_snapshot_when_capacity_allows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            paths = resolve_paths(Path(tmp), {})
            result = _create_bound_discovery(paths, pool_limit=12, successes=0)
            run_dir = Path(result.run_dir)
            for index in range(12):
                _record_demo_candidate(run_dir, suffix=f"approved-{index}")
            finish_current_keyword(run_dir)
            for profile in load_contact_index(run_dir, ACCOUNT):
                set_boss_decision(run_dir, ACCOUNT, profile, "发送")
            prepare_candidate_batch_review(run_dir)
            confirm_candidate_batch(run_dir)
            snapshot = prepare_dispatch_authorization(run_dir)
            payload = json.loads(next(Path(item) for item in snapshot.output_paths if Path(item).suffix == ".json").read_text(encoding="utf-8"))
            self.assertEqual(payload["approved_dispatch_count"], 12)
            self.assertEqual(payload["count"], 12)
            self.assertEqual(payload["interval_seconds"], 120)
            self.assertIn("完整贴文", snapshot.message)

    def test_feedback_is_cited_context_and_rule_needs_explicit_confirmation(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            run_dir = Path(tmp) / "reports" / "linkedin-search" / "run-1"
            run_dir.mkdir(parents=True)
            card = _card("feedback", "Complete feedback post")
            sync_candidate(run_dir, account_profile_url=ACCOUNT, card=card)
            set_boss_decision(run_dir, ACCOUNT, card["selected_member"]["profile_url"], "排除", "这是同类材料供应商")
            captured = capture_workbook_feedback(run_dir, ACCOUNT)
            feedback_id = captured["added_feedback_ids"][0]
            proposal = propose_screening_rule(
                run_dir,
                ACCOUNT,
                wording="明确自产同类基础材料的公司排除",
                scope="direct same-category manufacturers",
                supporting_feedback_ids=[feedback_id],
            )
            self.assertEqual(proposal["status"], "proposed_not_active")
            with self.assertRaisesRegex(ValueError, "明确确认"):
                confirm_screening_rule(run_dir, ACCOUNT, rule_id=proposal["rule_id"], confirmed=False)
            active = confirm_screening_rule(run_dir, ACCOUNT, rule_id=proposal["rule_id"], confirmed=True)
            self.assertEqual(active["status"], "confirmed_active")


def _create_bound_discovery(paths, *, pool_limit: int, successes: int):
    result = create_linkedin_search_run(
        paths,
        f"LinkedIn 搜索，关键词：exhaust wrap；排序最新；近一个月；不使用留言；间隔 2 分钟；最多找 {pool_limit} 人",
        now=datetime(2026, 7, 26, 9, 0, 0),
    )
    if successes:
        current = datetime(2026, 7, 26, 9, 1, 0).astimezone()
        path = ledger_path(Path(result.run_dir))
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "accounts": {
                        ACCOUNT: {
                            "contacts": {}, "companies": {}, "posts": {},
                            "dispatch_successes": [
                                {"candidate_id": f"old-{index}", "occurred_at": current.isoformat()}
                                for index in range(successes)
                            ],
                        }
                    },
                }
            ),
            encoding="utf-8",
        )
    bind_linkedin_account(
        Path(result.run_dir),
        LinkedInAccountObservation(True, "Tuolin Sales", ACCOUNT, True),
        browser_authorized=True,
    )
    record_first_posts_search(
        Path(result.run_dir),
        LinkedInPostSearchObservation("exhaust wrap", "posts", "latest", "past_month", 20, True, {}),
    )
    return result


def _card(suffix: str, post_text: str) -> dict:
    return {
        "candidate_id": f"candidate_{suffix}",
        "source_keyword": "exhaust wrap",
        "post_text": post_text,
        "post_url": f"https://linkedin.com/posts/{suffix}",
        "relevance_reason": "Visible business use may require high-temperature wrapping material.",
        "business_role": "downstream_material_user",
        "supporting_evidence": "The complete post describes exhaust insulation work.",
        "material_doubts": "No public RFQ is visible.",
        "company": {"name": f"Company {suffix}", "url": f"https://linkedin.com/company/{suffix}"},
        "selected_member": {
            "name": f"Person {suffix}",
            "title": "Founder",
            "company": f"Company {suffix}",
            "profile_url": f"https://linkedin.com/in/{suffix}",
        },
    }


if __name__ == "__main__":
    unittest.main()
